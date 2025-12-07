import math
import csv
import os
import sys
import tkinter as tk
from tkinter import messagebox
import xlsxwriter
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Basic constants for wheel size, motor speed and gravity
d = 0.645           # wheel diameter in m
rpm = 213           # rpm, later transfered to m/s
V_supply = 12       # volts
g = 9.81            # acceleration due to gravity in m/s^2

# Lever arm lengths for converting the hanging mass into effective tire load
l_hang = 0.875      # length hanging mass in m
l_reifen = 0.358    # length tire in m

# Column order used everywhere
FIELDNAMES = [
    "Tire name / type",
    "Tire pressure [bar]",
    "Idle currents [A]",
    "Load currents [A]",
    "Mean idle current [A]",
    "Mean load current [A]",
    "Weight on lever [kg]",
    "Effective weight on tire [kg]",
    "Speed [m/s]",
    "P_0 [W]",
    "P_load [W]",
    "P_rr [W]",
    "C_rr",
]


def get_unique_filename(base_path):
    """Return unique filename: base, base (2), base (3), ..."""
    if not os.path.exists(base_path):
        return base_path
    name, ext = os.path.splitext(base_path)
    counter = 2
    while True:
        candidate = f"{name} ({counter}){ext}"
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def parse_float_list(text):
    """Parse a space separated string of numbers into a list of floats."""
    parts = text.strip().split()
    if not parts:
        raise ValueError("No values entered")
    values = []
    for p in parts:
        p_clean = p.replace(",", ".")
        values.append(float(p_clean))
    return values


def format_value(key, value):
    """Format numeric and string values for CSV and Excel export."""
    if value is None or value == "":
        return ""

    if isinstance(value, (float, int)):
        decimal_places = {
            "Mean idle current [A]": 3,
            "Mean load current [A]": 3,
            "Speed [m/s]": 3,
            "P_0 [W]": 2,
            "P_load [W]": 2,
            "P_rr [W]": 2,
            "C_rr": 6,
            "Effective weight on tire [kg]": 3,
            "Weight on lever [kg]": 3,
            "Tire pressure [bar]": 2,
        }
        dp = decimal_places.get(key, 3)
        s = f"{float(value):.{dp}f}"
        return s.replace(",", ".")

    if isinstance(value, str):
        return value.replace(",", ".")

    return value


def compute_result(idle_values, load_values, m_hang):
    """Compute speed, currents, powers, effective weight and rolling resistance."""
    U = math.pi * d
    v = U * (rpm / 60.0)

    I_idle = sum(idle_values) / len(idle_values)
    I_load = sum(load_values) / len(load_values)

    m_eff = (m_hang * l_hang) / l_reifen

    P0 = V_supply * I_idle
    P_weighted = V_supply * I_load
    P_rr = P_weighted - P0
    C_rr = P_rr / (m_eff * g * v)

    return {
        "Speed [m/s]": v,
        "Mean idle current [A]": I_idle,
        "Mean load current [A]": I_load,
        "Weight on lever [kg]": m_hang,
        "Effective weight on tire [kg]": m_eff,
        "P_0 [W]": P0,
        "P_load [W]": P_weighted,
        "P_rr [W]": P_rr,
        "C_rr": C_rr,
    }


class RollingResistanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Rolling Resistance Calculator")

        # In memory list of measurement dictionaries for this run
        self.saved_rows = []
        # Last calculated result, used by Save to List
        self.last_result = None

        # Set file paths once per program run and ensure unique names
        script_dir = os.path.dirname(os.path.abspath(__file__))
        base_xlsx = os.path.join(script_dir, "rolling_resistance_data.xlsx")
        base_csv = os.path.join(script_dir, "rolling_resistance_data.csv")
        self.filename_xlsx = get_unique_filename(base_xlsx)
        self.filename_csv = get_unique_filename(base_csv)

        self._build_ui()

    def _build_ui(self):
        """Build all GUI elements and place them in the main window."""
        title_label = tk.Label(
            self.root,
            text="Rolling Resistance Calculator",
            font=("Arial", 24, "bold"),
            fg="steelblue"
        )
        title_label.grid(row=0, column=0, columnspan=2, pady=(10, 5))

        info_label = tk.Label(
            self.root,
            text=(
                "1.  Enter idle and load currents (use spaces between values).\n"
                "     The software will automatically calculate the mean of each set.\n"
                "     A test lasts 30 seconds and every 5 seconds a value should be added.\n"
                "2.  Enter the weight placed on the lever arm.\n"
                "     The software calculates the effective weight on the tire.\n"
                "3.  Enter tire name / type and tire pressure.\n"
                "     Changing the tire name / type creates a new line in the plot.\n"
                "4.  Use comma or dot for decimals (Excel export will convert to dots).\n\n"
                "-    Press Calculate to compute all values.\n"
                "-    Press Save to List to store the result.\n"
                "-    Press Go to Excel to export and open the table.\n"
                "-    Press Plot to visualize Crr vs tire pressure.\n"
                "-    If you delete a row in the Excel file and press Create Plot again\n"
                "     the plot will update correctly.\n"
                "-    If you restart the program it will not overwrite old files but create\n"
                "     new ones with increasing numbers."
            ),
            font=("Arial", 10, "bold"),
            justify="left"
        )
        info_label.grid(row=1, column=0, columnspan=2, pady=(0, 15))

        tk.Label(self.root, text="Tire name / type:", width=25, anchor="w").grid(row=2, column=0)
        self.entry_tire = tk.Entry(self.root, width=40)
        self.entry_tire.grid(row=2, column=1, padx=5, pady=2, sticky="w")

        tk.Label(self.root, text="Weight on lever in kg:", width=25, anchor="w").grid(row=3, column=0)
        self.entry_weight = tk.Entry(self.root, width=40)
        self.entry_weight.grid(row=3, column=1, padx=5, pady=2, sticky="w")

        tk.Label(self.root, text="Idle currents in A:", width=25, anchor="w").grid(row=4, column=0)
        self.entry_idle = tk.Entry(self.root, width=40)
        self.entry_idle.grid(row=4, column=1, padx=5, pady=2, sticky="w")

        tk.Label(self.root, text="Load currents in A:", width=25, anchor="w").grid(row=5, column=0)
        self.entry_load = tk.Entry(self.root, width=40)
        self.entry_load.grid(row=5, column=1, padx=5, pady=2, sticky="w")

        tk.Label(self.root, text="Tire pressure in bar:", width=25, anchor="w").grid(row=6, column=0)
        self.entry_pressure = tk.Entry(self.root, width=40)
        self.entry_pressure.grid(row=6, column=1, padx=5, pady=2, sticky="w")

        # Buttons row
        button_frame = tk.Frame(self.root)
        button_frame.grid(row=7, column=0, columnspan=2, pady=15)

        btn_calc = tk.Button(
            button_frame,
            text="Calculate",
            font=("Arial", 11, "bold"),
            command=self.on_calculate
        )
        btn_calc.pack(side="left", padx=5)

        btn_save = tk.Button(
            button_frame,
            text="Save to List",
            font=("Arial", 11, "bold"),
            command=self.on_save_to_list
        )
        btn_save.pack(side="left", padx=5)

        btn_excel = tk.Button(
            button_frame,
            text="Go to Excel",
            font=("Arial", 11, "bold"),
            command=self.on_go_to_excel
        )
        btn_excel.pack(side="left", padx=5)

        btn_plot = tk.Button(
            button_frame,
            text="Create Plot",
            font=("Arial", 11, "bold"),
            command=self.on_plot_pressure_vs_crr
        )
        btn_plot.pack(side="left", padx=5)

        # Output labels for computed values
        output_width = 20

        tk.Label(self.root, text="Speed v [m/s]:", width=25, anchor="w").grid(row=8, column=0)
        self.label_speed_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_speed_val.grid(row=8, column=1)

        tk.Label(self.root, text="Mean idle current [A]:", width=25, anchor="w").grid(row=9, column=0)
        self.label_I_idle_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_I_idle_val.grid(row=9, column=1)

        tk.Label(self.root, text="Mean load current [A]:", width=25, anchor="w").grid(row=10, column=0)
        self.label_I_load_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_I_load_val.grid(row=10, column=1)

        tk.Label(self.root, text="P_0 [W]:", width=25, anchor="w").grid(row=11, column=0)
        self.label_P0_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_P0_val.grid(row=11, column=1)

        tk.Label(self.root, text="P_load [W]:", width=25, anchor="w").grid(row=12, column=0)
        self.label_Pw_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_Pw_val.grid(row=12, column=1)

        tk.Label(self.root, text="P_rr [W]:", width=25, anchor="w").grid(row=13, column=0)
        self.label_Prr_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_Prr_val.grid(row=13, column=1)

        tk.Label(self.root, text="Weight on lever [kg]:", width=25, anchor="w").grid(row=14, column=0)
        self.label_mhang_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_mhang_val.grid(row=14, column=1)

        tk.Label(self.root, text="Effective weight on tire [kg]:", width=25, anchor="w").grid(row=15, column=0)
        self.label_meff_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_meff_val.grid(row=15, column=1)

        tk.Label(self.root, text="Tire name:", width=25, anchor="w").grid(row=16, column=0)
        self.label_tire_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_tire_val.grid(row=16, column=1)

        tk.Label(self.root, text="Pressure [bar]:", width=25, anchor="w").grid(row=17, column=0)
        self.label_pressure_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_pressure_val.grid(row=17, column=1)

        tk.Label(self.root, text="Rolling resistance C_rr:", width=25, anchor="w").grid(row=18, column=0)
        self.label_Crr_val = tk.Label(self.root, text="-", width=output_width, anchor="w")
        self.label_Crr_val.grid(row=18, column=1)

    def on_calculate(self):
        """Read inputs, run physics computation and update GUI labels."""
        try:
            idle_values = parse_float_list(self.entry_idle.get())
            load_values = parse_float_list(self.entry_load.get())

            w_text = self.entry_weight.get().replace(",", ".").strip()
            if not w_text:
                raise ValueError("No weight entered")
            m_hang = float(w_text)

            core_result = compute_result(idle_values, load_values, m_hang)

            tire_name = self.entry_tire.get().strip()
            tire_pressure = self.entry_pressure.get().replace(",", ".").strip()

            self.label_speed_val.config(text=f"{core_result['Speed [m/s]']:.3f}")
            self.label_I_idle_val.config(text=f"{core_result['Mean idle current [A]']:.3f}")
            self.label_I_load_val.config(text=f"{core_result['Mean load current [A]']:.3f}")
            self.label_P0_val.config(text=f"{core_result['P_0 [W]']:.3f}")
            self.label_Pw_val.config(text=f"{core_result['P_load [W]']:.3f}")
            self.label_Prr_val.config(text=f"{core_result['P_rr [W]']:.3f}")
            self.label_mhang_val.config(text=f"{core_result['Weight on lever [kg]']:.3f}")
            self.label_meff_val.config(text=f"{core_result['Effective weight on tire [kg]']:.3f}")
            self.label_Crr_val.config(text=f"{core_result['C_rr']:.6f}")
            self.label_tire_val.config(text=tire_name if tire_name else "-")
            self.label_pressure_val.config(text=tire_pressure if tire_pressure else "-")

            # Store last result as a dict so it can be exported or saved to list
            self.last_result = {
                "Tire name / type": tire_name,
                "Tire pressure [bar]": tire_pressure,
                "Idle currents [A]": self.entry_idle.get().strip(),
                "Load currents [A]": self.entry_load.get().strip(),
            }
            self.last_result.update(core_result)

        except ValueError as e:
            messagebox.showerror("Error", f"Input error: {e}")

    def _append_last_result_to_excel_if_exists(self):
        """Append the last result as a new row to the current Excel file if it exists."""
        if self.last_result is None:
            return

        filename_xlsx = self.filename_xlsx

        if not os.path.exists(filename_xlsx):
            return

        try:
            wb = load_workbook(filename_xlsx)
            if "Data" in wb.sheetnames:
                ws = wb["Data"]
            else:
                ws = wb.active

            # If the sheet is empty, write the header first
            if ws.max_row == 1 and all(c.value is None for c in ws[1]):
                for col, name in enumerate(FIELDNAMES, start=1):
                    ws.cell(row=1, column=col, value=name)

            new_row_idx = ws.max_row + 1
            # Use the first data row as a style template if available
            template_row_idx = 2 if ws.max_row >= 2 else 1

            for col, key in enumerate(FIELDNAMES, start=1):
                val = format_value(key, self.last_result.get(key, ""))
                new_cell = ws.cell(row=new_row_idx, column=col, value=val)

                template_cell = ws.cell(row=template_row_idx, column=col)
                if template_cell.has_style:
                    new_cell._style = template_cell._style

            wb.save(filename_xlsx)
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Could not append to Excel:\n{e}")

    def on_save_to_list(self):
        """Save the current result to the in memory list and to Excel if present."""
        if self.last_result is None:
            messagebox.showerror("Error", "Please press Calculate before Save to List")
            return
        self.saved_rows.append(self.last_result.copy())
        self._append_last_result_to_excel_if_exists()
        messagebox.showinfo("Saved", f"Data set number {len(self.saved_rows)} saved in list")

    def on_go_to_excel(self):
        """Create Excel if needed, sync CSV with Excel and open the Excel file."""
        filename_csv = self.filename_csv
        filename_xlsx = self.filename_xlsx

        try:
            # Create Excel file once from saved_rows if it does not exist yet
            if not os.path.exists(filename_xlsx):
                if not self.saved_rows:
                    messagebox.showerror("Error", "No data in list to export\nUse Save to List first")
                    return

                workbook = xlsxwriter.Workbook(filename_xlsx)
                worksheet = workbook.add_worksheet("Data")

                header_format = workbook.add_format({
                    "bold": True,
                    "bg_color": "#D9D9D9",
                    "align": "center",
                    "border": 1,
                })

                cell_format = workbook.add_format({
                    "align": "center",
                    "border": 1,
                })

                for col, name in enumerate(FIELDNAMES):
                    worksheet.write(0, col, name, header_format)

                for row_idx, row in enumerate(self.saved_rows, start=1):
                    for col_idx, key in enumerate(FIELDNAMES):
                        val = format_value(key, row.get(key, ""))
                        worksheet.write(row_idx, col_idx, val, cell_format)

                worksheet.set_column(0, len(FIELDNAMES) - 1, 18)
                workbook.close()

            # Read the current Excel content and regenerate the CSV so CSV follows Excel
            try:
                wb = load_workbook(filename_xlsx, data_only=True)
                if "Data" in wb.sheetnames:
                    ws = wb["Data"]
                else:
                    ws = wb.active

                header = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row is None:
                        continue
                    if all(cell is None for cell in row):
                        continue
                    row_dict = {}
                    for i, col_name in enumerate(header):
                        if col_name is None:
                            continue
                        val = row[i] if i < len(row) else ""
                        row_dict[str(col_name)] = val
                    rows.append(row_dict)
                wb.close()
            except Exception as e:
                messagebox.showerror("Error", f"Could not read Excel data:\n{e}")
                return

            # Write CSV from the Excel rows
            with open(filename_csv, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=FIELDNAMES, delimiter=";")
                writer.writeheader()
                for row in rows:
                    out = {key: format_value(key, row.get(key, "")) for key in FIELDNAMES}
                    writer.writerow(out)

            # Try to open the Excel file for the user
            try:
                if os.name == "nt":
                    os.startfile(filename_xlsx)
                elif sys.platform == "darwin":
                    os.system(f"open '{filename_xlsx}'")
                else:
                    os.system(f"xdg-open '{filename_xlsx}'")
            except Exception as e:
                messagebox.showerror("Error", f"File created but could not be opened:\n{e}")

        except Exception as e:
            messagebox.showerror("Error", f"Could not write file:\n{e}")

    def on_plot_pressure_vs_crr(self):
        """Read data from Excel or list and plot Crr versus tire pressure."""
        filename_xlsx = self.filename_xlsx

        rows = []

        # Prefer reading plot data from the Excel file so deletions there are respected
        if os.path.exists(filename_xlsx):
            try:
                wb = load_workbook(filename_xlsx, data_only=True)
                if "Data" in wb.sheetnames:
                    ws = wb["Data"]
                else:
                    ws = wb.active

                header = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row is None:
                        continue
                    if all(cell is None for cell in row):
                        continue
                    row_dict = {}
                    for i, col_name in enumerate(header):
                        if col_name is None:
                            continue
                        val = row[i] if i < len(row) else ""
                        row_dict[str(col_name)] = val
                    rows.append(row_dict)
                wb.close()
            except Exception as e:
                messagebox.showerror("Error", f"Could not read Excel data:\n{e}")
                return

        # Fallback to in memory saved_rows if Excel does not exist or has no data
        if not rows:
            if not self.saved_rows:
                messagebox.showerror("Error", "No data in list to plot\nUse Save to List first")
                return
            rows = self.saved_rows

        colors = [
            "steelblue",
            "firebrick",
            "seagreen",
            "darkorange",
            "purple",
            "gold",
            "black",
        ]

        # Group values by tire name so each tire has its own curve
        tire_groups = {}
        for row in rows:
            name = str(row.get("Tire name / type", "")).strip()
            if name == "":
                name = "Unknown tire"

            pressure_text = str(row.get("Tire pressure [bar]", "")).replace(",", ".")
            crr_value = row.get("C_rr", None)

            if pressure_text == "" or crr_value is None:
                continue

            try:
                p = float(pressure_text)
                crr = float(crr_value)
            except ValueError:
                continue

            if name not in tire_groups:
                tire_groups[name] = []

            tire_groups[name].append((p, crr))

        if not tire_groups:
            messagebox.showerror("Error", "No valid pressure and C_rr data to plot")
            return

        fig = plt.figure(figsize=(7, 4.3))
        ax = fig.add_subplot(111)

        # Plot curves for each tire type
        for i, (tire_name, values) in enumerate(tire_groups.items()):
            values_sorted = sorted(values, key=lambda x: x[0])
            pressures = [v[0] for v in values_sorted]
            crr_values = [v[1] for v in values_sorted]
            color = colors[i % len(colors)]

            ax.plot(
                pressures,
                crr_values,
                marker="o",
                markersize=5,
                linestyle="-",
                label=tire_name,
                color=color,
                picker=5,
            )

        ax.set_xlabel("Tire pressure / [bar]")
        ax.set_ylabel("Rolling resistance C_rr")
        ax.set_title("Rolling resistance vs Tire pressure", fontweight="bold")
        ax.grid(linestyle=":", alpha=0.7)
        ax.legend(title="Tire type", loc="upper right")
        ax.margins(x=0.05, y=0.05)

        fig.tight_layout()

        # Annotation object used to display x,y values on point click
        annot = ax.annotate(
            "",
            xy=(0, 0),
            xytext=(10, 10),
            textcoords="offset points",
            bbox=dict(boxstyle="round", fc="white", ec="black"),
            arrowprops=dict(arrowstyle="->"),
        )
        annot.set_visible(False)

        def on_pick(event):
            """Callback for matplotlib pick_event to show a small info box near the point."""
            ind = event.ind[0]
            xdata = event.artist.get_xdata()
            ydata = event.artist.get_ydata()
            x = xdata[ind]
            y = ydata[ind]

            dx = 10
            dy = 10

            x_min, x_max = ax.get_xlim()
            span_x = x_max - x_min
            if x > x_max - 0.07 * span_x:
                dx = -80

            y_min, y_max = ax.get_ylim()
            span_y = y_max - y_min
            if y > y_max - 0.07 * span_y:
                dy = -30

            offset = (dx, dy)

            annot.xy = (x, y)
            annot.set_position(offset)
            annot.set_text(f"p = {x:.2f} bar\nC_rr = {y:.6f}")
            annot.set_visible(True)
            fig.canvas.draw_idle()

        fig.canvas.mpl_connect("pick_event", on_pick)

        plt.show()


if __name__ == "__main__":
    root = tk.Tk()
    app = RollingResistanceApp(root)
    root.mainloop()
